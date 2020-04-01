#-*-coding:utf-8-*-
import openpyxl
import os
from utils import getcwd


#打开excel工作簿‪C:\\Users\\dell\\Desktop\\zhoubao---limeng.xlsx
wb = openpyxl.load_workbook('C:\\Users\\dell\\Desktop\\zhoubao---limeng.xlsx')
#当前活跃的表单
ws = wb.active
#获取所有表单名称
sheets = wb.sheetnames
print(sheets,type(sheets))

#创建一个表单
mysheet = wb.create_sheet('testsheet')
print(wb.sheetnames)

#获取指定表单


print('---'*10)
print(ws)
print(ws['A1'])#获取表格第一个对象
print(ws['A1'].value)

c = ws['A1']
print('Row{}, coullmu{}, Value{}'.format(c.row,c.column,c.value))


